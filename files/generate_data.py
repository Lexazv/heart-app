from itertools import zip_longest
from concurrent.futures import ThreadPoolExecutor

import openpyxl
import requests


# Specify filename with users data here:
USERS_DATA_FILE: str = 'files/data.xlsx'

# Specify if you want to get access tokens for created users.
GENERATE_TOKENS: bool = True

def prepare_users_data(filename: str) -> list:
    file = openpyxl.load_workbook(filename=filename).active
    cols = [
        "email", "password", "first_name", "last_name", "confirmed_password"
    ]

    users_data = [
        dict(
            zip_longest(
                cols, [cell.value for cell in row],
                fillvalue=file.cell(row=row_index + 2, column=2).value
            )
        ) for row_index, row in enumerate(file.iter_rows(min_row=2))
    ]

    return users_data


def fill_db_users(users_data: list) -> None:

    def _create_user(user_data: dict):
        response = requests.post(
            url='http://localhost:8000/api/users/registration', json=user_data
        )
        return response

    with ThreadPoolExecutor() as executor:
        responses = executor.map(_create_user, users_data)

    for response in responses:
        print(response.json())


def get_tokens(users_data: list) -> None:

    def _get_token(user_data: dict):
        response = requests.post(
            url='http://localhost:8000/api/users/login', json=user_data
        )
        return response

    tokens = {}

    for dataset in users_data:
        tokens[dataset["email"]]=_get_token(dataset).json()["token"]
    
    print(tokens)


if __name__ == '__main__':
    users_data = prepare_users_data(filename=USERS_DATA_FILE)
    fill_db_users(users_data=users_data)

    if GENERATE_TOKENS:
        get_tokens(users_data=users_data)
